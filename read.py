import openpyxl

wb=openpyxl.load_workbook('C:\codings\python\excel\data.xlsx')
wo110,cy110,cy111,ma110,uc100,cv110,cs110,cs111,psc,it112=0,0,0,0,0,0,0,0,0,0
sh1=wb[wb.sheetnames[0]]
for i in range(1,7):
    for j in range(1,11):
        if i>1:
            if sh1.cell(i,j).value=="wo110":
                wo110+=1
            elif sh1.cell(i,j).value=="cy110":
                cy110+=1
            elif sh1.cell(i,j).value=="cy111":
                cy111+=1
            elif sh1.cell(i,j).value=="ma110":
                ma110+=1
            elif sh1.cell(i,j).value=="uc100":
                uc100+=1
            elif sh1.cell(i,j).value=="cv110":
                cv110+=1
            elif sh1.cell(i,j).value=="cs110":
                cs110+=1
            elif sh1.cell(i,j).value=="cs111":
                cs111+=1
            elif sh1.cell(i,j).value=="psc":
                psc+=1
            elif sh1.cell(i,j).value=="it112":
                it112+=1
            #print(sh1.cell(i,j).value,end=" ")
        #if j==10:
            #print("\n")
course_names=["wo110","cy110","cy111","ma110","uc100","cv110","cs110","cs111","psc","it112"]
course_hours=[wo110*14,cy110*14,cy111*14,ma110*14,uc100*14,cv110*14,cs110*14,cs111*14,psc*14,it112*14]
#print(course_hours)
