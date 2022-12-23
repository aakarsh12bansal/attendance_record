from datetime import datetime
import openpyxl

wb=openpyxl.load_workbook('C:\codings\python\excel\Absent.xlsx')
wb1=openpyxl.load_workbook('C:\codings\python\excel\data.xlsx')
ws1=wb1[wb1.sheetnames[0]]
ws=wb[wb.sheetnames[0]]
wo110_a,cy110_a,cy111_a,ma110_a,uc100_a,cv110_a,cs110_a,cs111_a,psc_a,it112_a=0,0,0,0,0,0,0,0,0,0


# get current datetime
dt = datetime.now()
print('Datetime is:', dt)
day=dt.strftime('%A')
print(day)
list1=[]
for i in range(1,7):
    if ws1.cell(i,1).value==day:
        for j in range(1,11):
            list1.append(ws1.cell(i,j).value)
print(list1)

#Updating Absentee information
absent=[]
while True:
    a=input("Were you absent in any classes today? ")
    if a[0] in "Yy":
        n=int(input("how many subjects did you miss? "))
        for i in range(n):
            ctr=0
            x=input("enter course code of subject ")
            if x=="wo110" and x in list1:
                ws.cell(2,2).value=int(ws.cell(2,2).value)+list1.count(x)
            
            elif x=="cy110" and x in list1:
                ws.cell(2,3).value=int(ws.cell(2,3).value)+list1.count(x)
            
            elif x=="cy111" and x in list1:
                ws.cell(2,4).value=int(ws.cell(2,4).value)+list1.count(x)
            
            elif x=="ma110" and x in list1:
                ws.cell(2,5).value=int(ws.cell(2,5).value)+list1.count(x)
            
            elif x=="uc100" and x in list1:
                ws.cell(2,6).value=int(ws.cell(2,6).value)+list1.count(x)
            
            elif x=="cv110" and x in list1:
                ws.cell(2,7).value=int(ws.cell(2,7).value)+list1.count(x)
           
            elif x=="cs110" and x in list1:
                ws.cell(2,8).value=int(ws.cell(2,8).value)+list1.count(x)
            
            elif x=="cs111" and x in list1:
                ws.cell(2,9).value=int(ws.cell(2,9).value)+list1.count(x)
            
            elif x=="psc" and x in list1:
                ws.cell(2,10).value=int(ws.cell(2,10).value)+list1.count(x)
            
            elif x=="it112" and x in list1:
                ws.cell(2,11).value=int(ws.cell(2,11).value)+list1.count(x)
            
        break
    else:
        print("you are a good boy ")
        break
cs111_a=int(ws.cell(2,9).value)
cy110_a=int(ws.cell(2,3).value)
cy111_a=int(ws.cell(2,4).value)
ma110_a=int(ws.cell(2,5).value)
uc100_a=int(ws.cell(2,6).value)
cv110_a=int(ws.cell(2,7).value)
cs110_a=int(ws.cell(2,8).value)
wo110_a=int(ws.cell(2,2).value)
it112_a=int(ws.cell(2,11).value)
psc_a=int(ws.cell(2,10).value)


absent=[wo110_a,cy110_a,cy111_a,ma110_a,uc100_a,cv110_a,cs110_a,cs111_a,psc_a,it112_a]
print(absent)
wb.save('C:\codings\python\excel\Absent.xlsx')

import read
hours=read.course_hours
name=read.course_names
presence={}
for i in range(len(hours)):
    absence=(absent[i]/hours[i])*100
    absence=format(absence,".2f")
    presence[name[i]]=(absence)
print(presence)

for i,j in presence.items():
    if float(j)>20:
        print(f"Your attendance in {i} is low!")